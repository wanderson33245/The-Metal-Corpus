import os
import re
import openpyxl
from dotenv import load_dotenv
import time

load_dotenv()

vs_folder = os.environ.get("VS_FOLDER")
lyric_folder_name = "Lyrics"
lyric_folder_path = os.path.join(vs_folder, lyric_folder_name)
overnight_log_file_name = "overnight_log_clean.txt"

log_dir = os.path.join(vs_folder, "Logs")
os.makedirs(log_dir, exist_ok=True)
log_file = open(os.path.join(log_dir, overnight_log_file_name), "w")

start_time = time.time()

def log(message):
    print(message)
    log_file.write(message + "\n")
    log_file.flush()

def load_bands_to_process(vs_folder):
    xlsx_path = os.path.join(vs_folder, "bands_to_process.xlsx")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    bands = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        band_name = row[0]
        skip = row[5]
        if not band_name or skip == "YES":
            continue
        band_name = str(band_name) #openpyxl reads numeric Excel cells as integers. Bands like "1349" or "3" break that.
        bands.append({"band_name": band_name})
    return bands

def clean_filename(title):
    title = title.replace('&', 'and')
    title = title.replace('/', '_')
    title = re.sub(r'[<>:"/\\|?*]', '', title)
    return title.strip()

replacements = {
    "`": "'",
    "\u201c": "",  # left curly quote
    "\u201d": "",  # right curly quote
    "\u00ab": "",  # <<
    "\u00bb": "",  # >>
    "u\u0308": "u", "o\u0308": "o", "a\u0308": "a", "e\u0308": "e", "i\u0308": "i",
    "\u00fc": "u", "\u00f6": "o", "\u00e4": "a", "\u00eb": "e", "\u00ef": "i",
    "\u00e1": "a", "\u00e9": "e", "\u00ed": "i", "\u00f3": "o", "\u00fa": "u",
    "\u00f1": "n", "\u00e5": "a", "\u00f8": "o", "\u00e6": "ae"
}

def normalize_text(text):
    for accented, replacement in replacements.items():
        text = text.replace(accented, replacement)
    return text

def is_instrumental_only(lyrics):
    """Detect lyric sheets that are entirely instrumental descriptions with no real lyrics"""
    cleaned = lyrics.strip()
    cleaned = re.sub(r'^\d+ Contributor[s]?.*?Lyrics', '', cleaned, flags=re.DOTALL).strip()
    lines = [line.strip() for line in cleaned.split('\n') if line.strip()]
    if not lines:
        return True
    instrumental_patterns = [
        r'^instrumental.*$',
        r'^\(.*instrumental.*\)$',
        r'^-+.*instrumental.*-+$',
        r'^\[.*instrumental.*\]$',
        r'^\{.*instrumental.*\}$',
        r'^guitar solo$',
        r'^no lyrics$',
        r'^no lyrics available$',
    ]
    for line in lines:
        line_lower = line.lower()
        if not any(re.match(p, line_lower) for p in instrumental_patterns):
            return False
    return True

def expand_repeat_markers(lyrics):
    """Expand (x3) style repeat markers into repeated lines"""
    lines = lyrics.split('\n')
    output_lines = []
    for line in lines:
        match = re.search(r'\(x(\d+)\)?', line, re.IGNORECASE)
        if match:
            count = int(match.group(1))
            clean_line = re.sub(r'\s*\(x\d+\)?', '', line, flags=re.IGNORECASE).strip()
            if clean_line:
                for _ in range(count):
                    output_lines.append(clean_line)
        else:
            output_lines.append(line)
    return '\n'.join(output_lines)



def clean_words(lyrics):
    """Clean apostrophes at word boundaries"""
    lines = lyrics.split('\n')
    cleaned_lines = []
    for line in lines:
        words = line.split(' ')
        cleaned_words = []
        for word in words:
            if word.endswith("in'"):
                word = word[:-1] + "g"
            if word.startswith("'"):
                word = word[1:]
            if word.endswith("'"):
                word = word[:-1]
            cleaned_words.append(word)
        cleaned_lines.append(' '.join(cleaned_words))
    return '\n'.join(cleaned_lines)

def clean_lyrics(lyrics):

    # Strip invisible zero-width characters that can silently break every
    # downstream regex (seen in Genius scrapes splitting words like "1st" into
    # "1s\u200bt"). Must run before any other cleaning step.
    # This is because of The Anatomy Act of 1832 by Exhumed. The solos line had zero-width spaces for some reason.
    invisible_chars = '\u200b\u200c\u200d\u2060\ufeff'
    lyrics = lyrics.translate(str.maketrans('', '', invisible_chars))

    removed_first_line = False

    lines = lyrics.split('\n')

    # Handle first line
    if lines:
        first_line = lines[0]
        if "Contributor" in first_line and "Read More" in first_line:
            after_read_more = re.split(r'Read More', first_line, maxsplit=1)[-1].strip()
            if after_read_more:
                lines[0] = after_read_more
            else:
                lines = lines[1:]
            removed_first_line = True
        elif "Contributor" in first_line and "Lyrics" in first_line:
            after_lyrics = re.split(r'Lyrics', first_line, maxsplit=1)[-1].strip()
            if after_lyrics:
                lines[0] = after_lyrics
            else:
                lines = lines[1:]
            removed_first_line = True

    lyrics = '\n'.join(lines)
    lyrics = normalize_text(lyrics)
    lyrics = expand_repeat_markers(lyrics)

    # ================================================================
    # SECTION MARKERS — [], (), {} all bracket styles
    # ================================================================
    lyrics = re.sub(
        r'^\s*[\(\[\{](chorus|verse(?:\s*\d*)?|bridge|outro|intro|hook|pre-chorus|'
        r'interlude|instrumental|prechorus|post-chorus|postchorus|refrain|'
        r'pre-refrain|post-refrain|repeat chorus|repeat verse|repeat bridge|'
        r'instrumental break|instrumental outro|instrumental intro|'
        r'pre chorus|post chorus|cho)[^\)\]\}]*[\)\]\}]/?\s*$',
        '', lyrics, flags=re.IGNORECASE | re.MULTILINE
    )

    # Plain-text section markers
    section_words = r'(chorus|verse(?:\s*\d*)?|bridge|outro|intro|hook|pre-chorus|post-chorus|interlude|pre chorus|post chorus)'
    lyrics = re.sub(rf'^\s*(repeat\s+)?{section_words}\s*:?/?\s*$', '', lyrics, flags=re.IGNORECASE | re.MULTILINE)

    # Abbreviated: (cho:), (rep:), *chorus:, -Chorus-, standalone repeat/chorus
    lyrics = re.sub(r'^\s*[\(\[\{]?(cho|rep|pre|pch|ch)\s*:?\s*[\)\]\}]?\s*$', '', lyrics, flags=re.IGNORECASE | re.MULTILINE)
    lyrics = re.sub(r'^\s*-+\s*(chorus|verse|bridge|outro|intro|repeat|instrumental)[^\n]*-+\s*$', '', lyrics, flags=re.IGNORECASE | re.MULTILINE)
    lyrics = re.sub(r'^\s*\*\s*(chorus|verse|bridge|outro|intro)\s*:?\s*$', '', lyrics, flags=re.IGNORECASE | re.MULTILINE)
    lyrics = re.sub(r'^\s*(repeat|chorus)\s*$', '', lyrics, flags=re.IGNORECASE | re.MULTILINE)

    # ================================================================
    # INSTRUMENTAL MARKERS
    # ================================================================
    lyrics = re.sub(
        r'^\s*[\(\[\{-]*\s*(instrumental|guitar solo|bass solo|drum solo|'
        r'keyboard solo|piano solo|synth solo|acoustic guitar instrumental|'
        r'more instrumental|instrumental break|instrumental outro|'
        r'instrumental intro|short instrumental|dark ambient noise instrumental|'
        r'no lyrics)[^\n]*[\)\]\}-]*\s*$',
        '', lyrics, flags=re.IGNORECASE | re.MULTILINE
    )

    # Remove lines like "Thematic flute solos on this track by Cheryl Pyle" or "Acoustic lead by L. Strom"
    # (Both examples from A Retreat Into Delusions by Elvenking)

    credit_openers = r'lead|solo|solos|vocals|guitar|bass|drums|acoustic|thematic|backing|additional|sample|sung|flute|keys|piano'
    lyrics = re.sub(
        rf'^\s*(?:{credit_openers})[\w\s]*\bby\s+[A-Z][a-zA-Z\'\.]*(?:\s+[A-Z][a-zA-Z\'\.]*)*\s*$',
        '', lyrics, flags=re.MULTILINE | re.IGNORECASE
)

    # ================================================================
    # SOLO MARKERS
    # ================================================================

    # Bare solo marker, no performer named: Solo, SOLO, (solo), -Solo-, -Solos-, Guitar Solo
    lyrics = re.sub(
        r'^\s*[\(\[\{\-]*\s*(?:Guitar|Bass|Drum|Keyboard|Piano|Synth|Acoustic)?\s*Solos?\s*[\)\]\}\-]*\s*$',
        '', lyrics, flags=re.IGNORECASE | re.MULTILINE
    )

    # Solo credit naming one or more performers - handles an optional descriptor
    # prefix ("Trade-off solo:"), a vocal-range word instead of a name
    # ("Solo soprano:"), initials ("O.M"), filler words ("both"), and repeated
    # or comma-separated name lists:
    # Catches: (Solo: O.M, N.M), Trade-off solo: Kragen, Lee, Kragen, Lee,
    #          (Solos: Aydan, Jarpen, both), (Solo soprano:), Solo - Matthew Harvey
    name_token = r"[A-Z][a-zA-Z'\.]*|both|and"
    lyrics = re.sub(
        rf'^\s*[\(\[\{{]?\s*[A-Za-z\-]*\s*[Ss]olos?\s*(?:soprano|alto|tenor|bass|baritone|mezzo)?\s*[:\-]\s*'
        rf'(?:(?:{name_token})[\s,]*)*[\)\]\}}]?\s*$',
        '', lyrics, flags=re.MULTILINE
    )

    # Ordinal-numbered solo credits, whole line, whatever follows the ordinal
    # Catches: 
    # 1st and 2nd Solos – Michael Burke, 
    # 3rd and 4th Solos – Matthew Harvey
    # (See raw lyrics for The Anatomy Act of 1832 by Exhumed)
    lyrics = re.sub(
        r'^\s*\d+(?:st|nd|rd|th)(?:\s+and\s+\d+(?:st|nd|rd|th))?\s+[Ss]olos?\b.*$',
        '', lyrics, flags=re.IGNORECASE | re.MULTILINE
    )

    # ================================================================
    # SONGWRITING CREDITS
    # ================================================================
    lyrics = re.sub(
        r'^\s*[\(\[\{]?\s*(music|lyrics|lyrix|lyrixxx|words|music and lyrics|'
        r'lyrics and music|instrumentation|arranged by|written by|composed by|'
        r'lead|background lead|riffs|performed and arranged by|creation|'
        r'also performed on|words and music)\s*:?[^\n]*[\)\]\}]?\s*$',
        '', lyrics, flags=re.IGNORECASE | re.MULTILINE
    )
    # "(Ford, Ehmig)" style credits at start of song
    lyrics = re.sub(r'^\s*\([A-Z][a-z]+,\s*[A-Z][a-z]+\)\s*$', '', lyrics, flags=re.MULTILINE)

    # ================================================================
    # NAME / CHARACTER LABELS (guest vocalist credits, character dialogue tags)
    # ================================================================
    # All-caps names with colon: EGYPTIAN:, AYREON:, NARRATOR:
    lyrics = re.sub(r'^\s*[A-Z][A-Z\s]+:\s*$', '', lyrics, flags=re.MULTILINE)
    # With performer attribution: VILLAGERS (ROBERT SOETERBOEK):
    lyrics = re.sub(r'^\s*[A-Z][A-Z\s]+\s*\([^)]+\)\s*:\s*$', '', lyrics, flags=re.MULTILINE)
    # Single capitalized word in parentheses on own line: (Egyptian), (Hippie), (Pride)
    lyrics = re.sub(r'^\s*\([A-Z][a-z]+\)\s*$', '', lyrics, flags=re.MULTILINE)

    # Words that must NEVER be treated as a name label - common metal lyric
    # openers used as rhetorical address, e.g. "Love: a crime of severance"
    protected_label_words = {
        'life', 'world', 'eyes', 'love', 'death', 'night', 'blood', 'end',
        'nothing', 'day', 'pain', 'heart', 'fire', 'soul', 'hell', 'god',
        'time', 'hope', 'fear', 'faith', 'sorrow', 'silence', 'dreams',
        'rage', 'chaos', 'sin', 'shadows', 'misery', 'hate', 'destiny',
        'vengeance', 'glory', 'freedom', 'truth', 'lies', 'flesh', 'ashes',
        'darkness', 'sky', 'storm', 'war', 'chains', 'sun', 'moon', 'light', 'dark',
        'destination', 'behold'
    }

    def _strip_name_label(match):
        words = match.group(1).split()
        if any(w.lower() in protected_label_words for w in words):
            return match.group(0)  # looks like real lyric - leave untouched
        return ''  # strip the label

    # Generic 1-2 capitalized-word label before a colon: "Tom Englund: ",
    # "Burke: ", "Dr. Knox: ", "Chorus: " - guarded by the protected list above
    lyrics = re.sub(
        r'^\s*([A-Z][a-zA-Z\.\']*(?:\s+[A-Z][a-zA-Z\.\']*)?):\s+',
        _strip_name_label, lyrics, flags=re.MULTILINE
    )

   # Speaker attribution in parens, one or more names (including lowercase name
    # particles like "van", "von", "de", and capitalized middle initials),
    # may be followed by lyric text on the same line:
    # Catches: 
    # (Ty Tabor:) We all know, (Connect the Dots by Ayreon)
    # (Anneke van Giersbergen:) No need to feel desire (Comatose by Ayreon)
    name_particles = r'van|von|de|der|den|del|di|la|le|da|dos|das|el|al'
    lyrics = re.sub(
        rf'^\s*\((?:(?:[A-Z][a-zA-Z\'\.]*|{name_particles})\s*)+:\)\s*',
        '', lyrics, flags=re.MULTILINE
    )

    # Metal is pretty politically charged; they'll often include speeches by important 
    # historical figures, so this will get rid of any parenthesis that appear
    # afterward and credit the figure.
    # Example: Unnatural Selection by Ayreon
    quotable_people = [
        "JFK", "John F\.? Kennedy", "Al Gore", "George Bush", "George W\.? Bush",
        "Franklin Delano Roosevelt", "FDR", "Winston Churchill",
        "Abraham Lincoln", "Martin Luther King(?:,? Jr\.?)?", "Mahatma Gandhi",
        "Adolf Hitler", "Joseph Stalin", "Nietzsche", "Albert Einstein",
        "Ronald Reagan", "Barack Obama", "Napoleon", "Mussolini",
        "Malcolm X", "Nelson Mandela", "Sun Tzu", "Donald Trump",
    ]
    people_pattern = "|".join(quotable_people)
    lyrics = re.sub(
        rf'\s*\((?:{people_pattern})\)\s*$',
        '', lyrics, flags=re.MULTILINE | re.IGNORECASE
    )

    # "Solo" stage directions naming a performer, whole line: Solo – Matthew Harvey
    lyrics = re.sub(
        r'^\s*Solo\s*[-–—]\s*[A-Z][a-zA-Z\.\']*(?:\s+[A-Z][a-zA-Z\.\']*)*\s*$',
        '', lyrics, flags=re.MULTILINE | re.IGNORECASE
    )

    # ================================================================
    # ROMAN NUMERAL / LETTER / CHAPTER / PART SECTION HEADERS
    # ================================================================
    lyrics = re.sub(r'^\s*[IVX]+\.\s+.*$', '', lyrics, flags=re.MULTILINE)
    lyrics = re.sub(r'^\s*[A-Z]\.\s+.*$', '', lyrics, flags=re.MULTILINE)
    lyrics = re.sub(r'^\s*(Chapter|Part)\s+[IVX\d]+[:\.]?\s*.*$', '', lyrics, flags=re.IGNORECASE | re.MULTILINE)

    # ================================================================
    # TIMESTAMP ANNOTATIONS
    # ================================================================
    lyrics = re.sub(r'\(\d+:\d+\s*[-]\s*\d+:\d+\)', '', lyrics)

    # ================================================================
    # VERSION LABELS
    # ================================================================
    lyrics = re.sub(r'^\s*\(version\s*\d+\)\s*$', '', lyrics, flags=re.IGNORECASE | re.MULTILINE)

    # ================================================================
    # MISCELLANEOUS
    # ================================================================
    lyrics = re.sub(r'\betc\.?\b', '', lyrics, flags=re.IGNORECASE) #get rid of all "etc"
    lyrics = re.sub(r'_{2,}', ' ', lyrics)
    lyrics = re.sub(r'\*[^*]*\*', '', lyrics)
    lyrics = re.sub(r'[^\n]*\u2026\s*Read More[^\n]*\n?', '', lyrics)
    
    # Skit directions and vocal/action descriptors in parentheses
    # Catches: (laughs evilly), (Davey whines again), (screaming), (baby crying)
    action_verbs = (
        r'laughs?|cry|cries|whines?|screams?|moans?|groans?|sighs?|whispers?|shouts?|yells?|'
        r'sobs?|chuckles?|growls?|roars?|wails?|shrieks?|gasps?|breathes?|coughs?|grunts?|'
        r'mumbles?|mutters?|speaks?|spoken|talking|fading|fades?|vocalizes?|vocaliz|'
        r'instrumental|claps?|applauds?|noise|sounds?|echo(?:es)?|coos?|hums?|speaks?|'
        r'talks?'
    )
    lyrics = re.sub(
        rf'\s*\([^)]*({action_verbs})[^)]*\)\s*$',
        '', lyrics, flags=re.IGNORECASE | re.MULTILINE
    )

    # Mid-lyric annotation triggers
    annotation_triggers = [
        r'According to',
        r'From an interview',
        r'In a \d{4}',
        r'Chuck revealed',
        r'An acclaimed',
        r'The eponymous',
        r'This song (is|was|describes|details|discusses|paints|shows)',
        r'"[^"]+" (is|was|shows|expresses|paints|describes|emphasizes)',
        r'The whole album',
        r'The journey begins',
        r'The (first|second|third|fourth|fifth|final|last|opening|closing) (track|song)',
        r'\d+(st|nd|rd|th) (track|song)',
    ]
    for trigger in annotation_triggers:
        lyrics = re.sub(rf'[^\n]*{trigger}[^\n]*\n?', '', lyrics)

    lyrics = re.sub(r'\n{3,}', '\n\n', lyrics)
    lyrics = lyrics.strip()

    # If first remaining line is long prose, remove it
    lines = [line for line in lyrics.split('\n') if line.strip()]
    if len(lines) > 1 and len(lines[0].split()) > 15:
        lines = lines[1:]
        lyrics = '\n'.join(lines)
        removed_first_line = True

    lyrics = clean_words(lyrics)

    return lyrics.strip(), removed_first_line

def has_repeat_markers(lyrics):
    return bool(re.search(r'(x\d+|\(x\d+\)?)', lyrics, re.IGNORECASE))

def has_section_labels_without_breaks(lyrics):
    """Detect songs with section labels but no blank line delimiters"""
    section_pattern = re.compile(
        r'(?<!\n\n)\n\s*(repeat\s+)?(chorus|verse(?:\s*\d*)?|bridge|outro|intro|hook|pre-chorus|interlude)\s*:?\s*\n',
        re.IGNORECASE
    )
    return bool(section_pattern.search(lyrics))

def clean_band_lyrics(band_name, lyrics_folder):
    band_path = os.path.join(lyrics_folder, band_name)
    raw_path = os.path.join(band_path, "raw")
    xlsx_path = os.path.join(lyrics_folder, band_name, f"{band_name}_preview.xlsx")

    if not os.path.exists(raw_path):
        log(f"No raw folder found for {band_name}, skipping")
        return

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    for release_folder in os.listdir(raw_path):
        release_raw_path = os.path.join(raw_path, release_folder)
        if not os.path.isdir(release_raw_path):
            continue

        release_clean_path = os.path.join(band_path, "clean", release_folder)

        for filename in os.listdir(release_raw_path):
            if not filename.endswith(".txt"):
                continue

            raw_txt_path = os.path.join(release_raw_path, filename)
            clean_txt_path = os.path.join(release_clean_path, filename)

            if os.path.exists(clean_txt_path):
                log(f"  [already cleaned, skipping] {filename}")
                continue

            with open(raw_txt_path, "r") as f:
                raw_lyrics = f.read()

            # Check if entirely instrumental
            if is_instrumental_only(raw_lyrics):
                log(f"  [instrumental only, skipping] {filename}")
                status = "instrumental only"
                track_title_to_match = filename[:-4]
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if row[6] and clean_filename(str(row[6])) == track_title_to_match:
                        ws.cell(row=i, column=10).value = status
                        break
                continue

            # Check for section labels without breaks
            if has_section_labels_without_breaks(raw_lyrics):
                log(f"  [needs manual review - no section breaks] {filename}")
                status = "needs manual review - no section breaks"
                track_title_to_match = filename[:-4]
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if row[6] and clean_filename(str(row[6])) == track_title_to_match:
                        ws.cell(row=i, column=10).value = status
                        break
                continue

            cleaned, removed_first_line = clean_lyrics(raw_lyrics)

            if not cleaned.strip():
                log(f"  [empty after cleaning, skipping] {filename}")
                status = "empty after cleaning"
            else:
                os.makedirs(release_clean_path, exist_ok=True)
                with open(clean_txt_path, "w") as f:
                    f.write(cleaned)

                status = "cleaned"
                if removed_first_line:
                    status += ". Removed first line"
                if has_repeat_markers(cleaned):
                    status += ". Has repeat markers"
                log(f"  {status}: {filename}")

            track_title_to_match = filename[:-4]
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[6] and clean_filename(str(row[6])) == track_title_to_match:
                    ws.cell(row=i, column=10).value = status
                    break

    wb.save(xlsx_path)
    log(f"Saved updated xlsx for {band_name}")


bands_to_process = load_bands_to_process(vs_folder)

for band in bands_to_process:
    log(f"\n{'='*50}")
    log(f"Cleaning: {band['band_name']}")
    log(f"{'='*50}")
    clean_band_lyrics(band['band_name'], lyric_folder_path)



elapsed = time.time() - start_time
minutes = int(elapsed // 60)
seconds = int(elapsed % 60)
print(f"Total runtime: {minutes}m {seconds}s")