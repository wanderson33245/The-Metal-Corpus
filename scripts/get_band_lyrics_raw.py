import os
import re
import json
import time
import openpyxl
import lyricsgenius
import difflib
from dotenv import load_dotenv

load_dotenv()

genius = lyricsgenius.Genius(os.environ.get("GENIUS_TOKEN"))
genius.verbose = False
genius.remove_section_headers = True

vs_folder = os.environ.get("VS_FOLDER")
lyric_folder_name = "Lyrics"
lyric_folder_path = os.path.join(vs_folder, lyric_folder_name)
overnight_log_file_name = "overnight_log_raw.txt"

log_dir = os.path.join(vs_folder, "Logs")
os.makedirs(log_dir, exist_ok=True)
log_file = open(os.path.join(log_dir, overnight_log_file_name), "w")

def log(message):
    print(message)
    log_file.write(message + "\n")
    log_file.flush()

def clean_filename(title):
    title = title.replace('&', 'and')
    title = title.replace('/', '_') # Replaces "/" with "_". Death's "Rehearsal 28/3/86" was messing this up
    title = re.sub(r'[<>:"/\\|?*]', '', title) # removes other problematic characters
    return title.strip()

def is_valid_song(title):
    title_lower = title.lower()
    invalid_patterns = ['(live)', '- live', '(demo', '- demo', '(acoustic)', '- acoustic',
                        '(remix)', '- remix', '(remaster)', '- remaster', '(instrumental)',
                        '- instrumental', 'cover)', '- cover', 'demo version']
    return not any(pattern in title_lower for pattern in invalid_patterns)

def is_valid_retrieved_song(title):
    #Check that the song Genius returned isn't a live, demo, or alternate version
    title_lower = title.lower()
    invalid_endings = [
        'demo)', 'alternate)', 'version)', 'mix)', 'remix)', 'edit)',
        'remaster)', 'remastered)', 'acoustic)', 'instrumental)',
        'live)', 'take)', 'session)', 'outtake)', 'bonus)'
    ]
    invalid_contains = ['live at', 'live in', 'live from']
    if any(title_lower.endswith(ending) for ending in invalid_endings):
        return False
    if any(phrase in title_lower for phrase in invalid_contains):
        return False
    return True

def is_correct_artist(returned_artist, expected_artist, threshold=0.6):
    ratio = difflib.SequenceMatcher(None, returned_artist.lower(), expected_artist.lower()).ratio()
    return ratio >= threshold

def build_json_from_xlsx(band_name, lyrics_folder, country, vocalist_gender):
    wb = openpyxl.load_workbook(os.path.join(lyrics_folder, band_name, f"{band_name}_preview.xlsx"))
    ws = wb.active

    releases = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        include, release_title, year, release_type, subgenres, row_vocalist_gender, track_title, track_title_clean, get_lyrics_status, clean_status, song_retrieved, band_retrieved_from = row

        if include != "YES":
            continue

        release_title = clean_filename(release_title)

        if release_title not in releases:
            releases[release_title] = {
                "year": year,
                "type": release_type,
                "subgenres": [s.strip() for s in subgenres.split("|")] if subgenres else ["other"],
                "vocalist_gender": row_vocalist_gender
            }

    band_data = {
        "country": country,
        "vocalist_gender": vocalist_gender,
        "releases": releases
    }

    json_path = os.path.join(lyrics_folder, band_name, f"{band_name}.json")
    with open(json_path, "w") as f:
        json.dump(band_data, f, indent=4)
    log(f"Saved {json_path}")

def get_band_lyrics_raw(band_name, lyrics_folder, country, vocalist_gender, genius_name=None):
    genius_name = genius_name or band_name
    xlsx_path = os.path.join(lyrics_folder, band_name, f"{band_name}_preview.xlsx")

    if not os.path.exists(xlsx_path):
        log(f"No preview xlsx found for {band_name}, skipping")
        return

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    seen_songs = set()
    received_songs = set()
    band_path = os.path.join(lyrics_folder, band_name)
    os.makedirs(band_path, exist_ok=True)

    current_release = None
    release_path = None

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        include, release_title, year, release_type, subgenres, row_vocalist_gender, track_title, track_title_clean, get_lyrics_status, clean_status, song_retrieved, band_retrieved_from = row

        if include != "YES":
            ws.cell(row=i, column=9).value = "excluded"
            continue

        if not track_title:
            continue

        if release_title != current_release:
            current_release = release_title
            release_path = os.path.join(band_path, "raw", clean_filename(release_title))

        if not is_valid_song(track_title):
            log(f"  [filtered, skipping] {track_title}")
            ws.cell(row=i, column=9).value = "filtered"
            continue

        if track_title.lower() in seen_songs:
            log(f"  [duplicate, skipping] {track_title}")
            ws.cell(row=i, column=9).value = "duplicate"
            continue
        seen_songs.add(track_title.lower())

        txt_path = os.path.join(release_path, f"{clean_filename(track_title)}.txt")

        if os.path.exists(txt_path):
            log(f"  [already exists] {track_title}")
            ws.cell(row=i, column=9).value = "already exists"
            continue

        log(f"  Fetching: {track_title}")
        try:
            song = genius.search_song(track_title, genius_name)
        except Exception as e:
            log(f"  [error: {e}, retrying in 3 seconds...] {track_title}")
            time.sleep(3)
            try:
                song = genius.search_song(track_title, genius_name)
            except Exception as e:
                log(f"  [failed after retry, skipping] {track_title}")
                ws.cell(row=i, column=9).value = "timeout"
                continue

        if not song and "(" in track_title:
            simplified_title = track_title.split("(")[0].strip()
            log(f"  [not found, trying simplified title] {simplified_title}")
            try:
                song = genius.search_song(simplified_title, genius_name)
            except Exception as e:
                log(f"  [timeout on simplified title, skipping] {simplified_title}")
                song = None

        if song and not is_valid_retrieved_song(song.title):
            log(f"  [retrieved song is a live/demo/alternate version, skipping] {song.title}")
            ws.cell(row=i, column=9).value = f"invalid retrieved version: {song.title}"
            ws.cell(row=i, column=11).value = song.title
            ws.cell(row=i, column=12).value = song.artist
            continue

        if song and song.title.strip().lower() in received_songs:
            log(f"  [duplicate-received song, skipping] {song.title}")
            ws.cell(row=i, column=9).value = "duplicate received song"
            continue

        if song:
            if not is_correct_artist(song.artist, genius_name):
                log(f"  [wrong artist: got '{song.artist}', expected '{genius_name}', skipping]")
                ws.cell(row=i, column=9).value = f"wrong artist: {song.artist}"
                ws.cell(row=i, column=12).value = song.artist
                continue

            received_songs.add(song.title.strip().lower())
            if not os.path.exists(release_path):
                log(f"  Release validated. Creating new folder: {release_title}")
                os.makedirs(release_path, exist_ok=True)
            if song.lyrics and song.lyrics.strip():
                with open(txt_path, "w") as f:
                    f.write(song.lyrics)
                log(f"  Saved raw: {track_title}")
                ws.cell(row=i, column=9).value = "raw saved"
                ws.cell(row=i, column=11).value = song.title
                ws.cell(row=i, column=12).value = song.artist
            else:
                log(f"  [empty lyrics, skipping] {track_title}")
                ws.cell(row=i, column=9).value = "empty"
        else:
            log(f"  [not found on Genius] {track_title}")
            ws.cell(row=i, column=9).value = "not found"

    wb.save(xlsx_path)
    log("Creating JSON file...")
    build_json_from_xlsx(band_name, lyrics_folder, country, vocalist_gender)


def load_bands_to_process(vs_folder):
    xlsx_path = os.path.join(vs_folder, "bands_to_process.xlsx")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    bands = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        band_name, country, vocalist_gender, genius_name, discogs_name, skip, OG_order, validity = row
        if not band_name or skip == "YES":
            continue
        bands.append({
            "band_name": band_name,
            "country": country,
            "vocalist_gender": vocalist_gender or "male",
            "genius_name": genius_name or band_name,
            "discogs_name": discogs_name or band_name
        })
    return bands

bands_to_process = load_bands_to_process(vs_folder)

for band in bands_to_process:
    log(f"\n{'='*50}")
    log(f"Processing: {band['band_name']}")
    log(f"{'='*50}")
    if len(bands_to_process) > 5: # If it's fewer than 6, it... 1. is probably just a test run, and 2. Most likely won't overload Genius' API
        time.sleep(20) # A precaution agaionst overloading Genius' API
    get_band_lyrics_raw(
        band["band_name"],
        lyric_folder_path,
        band["country"],
        band["vocalist_gender"],
        genius_name=band["genius_name"]
    )

print("Done with this batch")
