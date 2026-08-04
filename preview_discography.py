import os
import time
import re
import json
import difflib
import openpyxl
import discogs_client
import musicbrainzngs
from dotenv import load_dotenv

load_dotenv()

discogs = discogs_client.Client('HeavyMetalWords/1.0', os.environ.get("DISCOGS_TOKEN"))
musicbrainzngs.set_useragent("HeavyMetalWords", "1.0", os.environ.get("MY_EMAIL"))

vs_folder = os.environ.get("VS_FOLDER")
lyric_folder_name = "Lyrics"
lyric_folder_path = os.path.join(vs_folder, lyric_folder_name)

def is_valid_release(release_data):
    invalid_formats = ['live', 'compilation', 'box set', 'mixtape', 'interview', 'video', 'dvd']
    invalid_title_keywords = [
        'box set', 'collection', 'greatest hit', 'best of', 'archive', 'archives', 'b-sides', '(demo)', '(demos)', 'radio selection',
        'tapes', 'sampler', 'mania', 'profile', 'video', 'press pack', 'souvenir', 'interview', 'wasted years', 'early days', 'history of',
        'visions of', 'over hammersmith', 'extraits', 'the essential', 'ten years', '10 years', 'twenty years', '20 years', 'thirty years',
        '30 years', 'fourty years', '40 years', 'fifty years', '50 years', 'decade of', 'decades of'
    ] #Dream Theater's Greatest Hits album is titled "Greatest Hit", singular, so this will catch that still

    release_format = release_data.get('format', '').lower()
    title_lower = release_data.get('title', '').lower()

    if any(keyword in release_format for keyword in invalid_formats):
        print(f"{title_lower} is excluded at is valid release format")
        return "excluded"
    if any(keyword in title_lower for keyword in invalid_title_keywords):
        print(f"{title_lower} is excluded at is valid release title keywords")
        return "excluded"

    # Catch "live" without catching "alive", "lives", "lived", "liver" (Like the word "deliver" or "livery")
    if re.search(r'live', title_lower) and not re.search(r'alive|lives\b|lived\b|liver\b', title_lower):
        print(f"{title_lower} is excluded at is valid release live")
        return "flagged"
    return "valid"

def search_mb_artist(band_name, limit=10):
    results = musicbrainzngs.search_artists(artist=band_name, limit=limit)
    artist_list = results["artist-list"]

    for artist in artist_list:
        if artist.get("name", "").lower() == band_name.lower(): #Prefer exact matches. Originally, searching for Death would bring up Napalm Death
            return artist

    # No exact match. Only fall back if the top result is actually similar. I kept getting The Beatles whenever a small band had "The" in their name and it takes multiple minutes to fetch The Beatles' discography
    if artist_list:
        top_result = artist_list[0]
        top_name = top_result.get("name", "")
        similarity = difflib.SequenceMatcher(None, band_name.lower(), top_name.lower()).ratio()
        if similarity >= 0.6:
            print(f"  [no exact match for '{band_name}', using top result: {top_name}] (similarity: {similarity:.2f})")
            return top_result
        else:
            print(f"  [no exact match for '{band_name}', top result '{top_name}' too dissimilar ({similarity:.2f}), skipping MusicBrainz]")
            return None

    return None

def get_mb_release_types(band_name):
    artist = search_mb_artist(band_name)
    if not artist:
        print(f"No MusicBrainz artist found for {band_name}")
        return {}

    artist_id = artist["id"]
    print(f"Fetching every release for {band_name}, MusicBrainz artist ID: {artist_id}")

    type_lookup = {}
    offset = 0
    seen_releases = set()

    while True:
        result = musicbrainzngs.browse_release_groups(artist=artist_id, limit=25, offset=offset)
        rg_list = result["release-group-list"]
        print(f"Fetched {len(rg_list)} release groups at offset {offset}")

        for rg in rg_list:
            title = rg.get("title", "").rstrip("…").strip().lower() #"History Repeats..." by Dying Fetus broke this
            mb_type = rg.get("type", "other")
            if title in seen_releases:
                print(f"  {mb_type} - EXCLUDED (duplicate) - {title}")
                continue
            print(f"  {mb_type} - {title}")
            type_lookup[title] = mb_type
            seen_releases.add(title)

        if len(rg_list) < 25:
            break

        offset += 25

    return type_lookup

def get_release_type_fuzzy(album_title, band_release_types, threshold=0.8):
    album_lower = album_title.lower()
    type_priority = {"Album": 0, "EP": 1, "Single": 2, "Compilation": 3, "Other": 4, "Unknown": 5}

    # Try exact match first
    if album_lower in band_release_types:
        return band_release_types[album_lower]

    # Fall back to fuzzy match
    best_match = None
    best_ratio = 0
    best_priority = 99

    for mb_title, mb_type in band_release_types.items():
        ratio = difflib.SequenceMatcher(None, album_lower, mb_title).ratio()
        priority = type_priority.get(mb_type, 99)
        # Prefer higher ratio; break ties by type priority (Album > EP > Single etc.)
        if ratio > best_ratio or (ratio == best_ratio and priority < best_priority):
            best_ratio = ratio
            best_match = mb_title
            best_priority = priority

    if best_ratio >= threshold and best_match:
        print(f"  [fuzzy match] '{album_title}' -> '{best_match}' ({best_ratio:.2f})")
        return band_release_types[best_match]

    print(f"  [no match found for] '{album_title}'")
    return "Unknown"

def clean_track_title(title):
    title = title.lower()
    title = re.sub(r'[^\w\s]', '', title)  # Removes punctuation: parentheses, apostrophes, dashes, etc.
    title = re.sub(r'\s+', ' ', title).strip()  # Collapses extra whitespace
    return title

def preview_discography(band_name, vocalist_gender, discogs_name=None):
    
    band_path = os.path.join(lyric_folder_path, band_name)
    xlsx_path = os.path.join(band_path, f"{band_name}_preview.xlsx")
    if os.path.exists(xlsx_path):
        print(f"Preview already exists for {band_name}, skipping")
        return
    
    discogs_name = discogs_name or band_name
    results = discogs.search(discogs_name, type='artist')
    if len(results) == 0:
        print(f"No Discogs artist found for '{discogs_name}' or they don't have any releases on there. Make sure you spelled it correctly")
        return
    discogs_artist = results[0]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Discography"
    ws.append(["include", "release_title", "year", "type", "subgenres", "vocalist_gender", "track_title", "track_title_clean", "get_lyrics_status", "clean_status", "song_retrieved", "band_retrieved_from"])

    print(f"Number of releases: {len(discogs_artist.releases)}")

    all_releases = []
    if len(discogs_artist.releases) <= 50:
        # Just fetch page 1, no need to paginate
        all_releases = list(discogs_artist.releases.page(1))
        print(f"[Discogs] Fetched page 1 only ({len(all_releases)} releases)")
    else:
        page = 1
        while True:
            time.sleep(5)
            try:
                releases_page = discogs_artist.releases.page(page)
                if not releases_page:
                    break
                all_releases.extend(releases_page)
                print(f"[Discogs] Fetched page {page}, {len(releases_page)} releases")
                page += 1
                time.sleep(2)
            except discogs_client.exceptions.HTTPError as e:
                if '429' in str(e):
                    print(f"  [Discogs][rate limited on page {page}, waiting 60 seconds]")
                    time.sleep(60)
                else:
                    print(f"  [Discogs error on page {page}, stopping] {e}")
                    break
            except Exception as e:
                print(f"  [Discogs error on page {page}, stopping] {e}")
                break

    included_releases = []

    for release in all_releases:
        data = release.data

        if data.get('type') != 'master':
            continue

        if data.get('role') != 'Main':
            continue

        validity = is_valid_release(data) # can be valid, flagged, or excluded. valid is totally fine like The Number of the Beast by Iron Maiden, flagged is "This might be a live album", so it could be Live '06 by Teenage Bottlerocket or Long Live Heavy Metal by 3 Inches of Blood. Excluded is just obviously something that doesn't need to be included.

        if validity == "excluded":
            continue

        title = data['title'].lower()

        if (
            title.endswith(('live', 'live)', 'sampler', 'sampler)'))
            or 'live at' in title
            or 'live in' in title
        ):
            #print(f"{data.get('title')} is excluded at title endswith live") #debug line. keep
            continue

        include_value = "YES" if validity == "valid" else ""
        included_releases.append({"data": data, "include": include_value})
        print(f"{'Including' if include_value else 'Flagging'}: {data.get('title')}")

    if len(included_releases) == 0:
        print("No releases validated. Fetching the first 10 releases that clearly aren't live albums...")
        count = 0
        for release in discogs_artist.releases:
            if count >= 10:
                break
            data = release.data
            title = data['title'].lower()
            if (
                title.endswith(('live', 'live)', 'sampler', 'sampler)'))
                or 'live at' in title
                or 'live in' in title
            ):
                continue
            included_releases.append({"data": data, "include": "YES"})
            print(f"Including: {data.get('title')}")
            count += 1

    # ----- Sort releases by albums, then EPs, then singles, then anything else
    band_release_types = get_mb_release_types(band_name)
    type_order = {"Album": 0, "EP": 1, "Single": 2, "Compilation": 3, "Other": 4, "Unknown": 5}

    def get_release_priority(release):
        return type_order.get(get_release_type_fuzzy(release["data"]["title"], band_release_types), 5)

    included_releases.sort(key=get_release_priority)

    seen_songs = set()

    use_delay = (len(included_releases) > 50)
    print(f"{len(included_releases)} releases to fetch, {'using delay' if use_delay else 'no delay needed'}")

    for release in included_releases:

        if use_delay:
            time.sleep(1.5)

        album_title = release["data"]["title"]
        include_value = release["include"]
        release_type = get_release_type_fuzzy(album_title, band_release_types)
        print(f"checking release {album_title}")

        try:
            time.sleep(4)
            master = discogs.master(release["data"]["id"])
            master.refresh()
            styles = master.styles or []
        except discogs_client.exceptions.HTTPError as e:
            if hasattr(e, 'response'):
                print(e.response.headers)
            if '429' in str(e):
                print(f"  [rate limited, waiting 60 seconds]")
                time.sleep(60)
                try:
                    master.refresh()
                    styles = master.styles or []
                except Exception as e2:
                    print(f"  [failed after wait, skipping] {album_title}")
                    continue
            else:
                print(f"  [Discogs error for {album_title}, skipping] {e}")
                continue
        except json.JSONDecodeError:
            try:
                master.refresh()
                styles = master.styles or []
            except Exception as e:
                print(f"  [failed after wait, skipping] {album_title}")
                continue
        except Exception as e:
            print(f"  [Discogs error for {album_title}, skipping] {e}")
            continue

        print(f"{album_title} made it through. Retrieving its data...")
        year = release["data"].get('year', 'Unknown')
        subgenres = ' | '.join(styles) if styles else "other"

        new_tracks = []
        for track in master.tracklist:
            if not track.title:
                print(f"  [no title, skipping]")
                continue
            if clean_track_title(track.title) not in seen_songs:
                new_tracks.append(track.title.strip())

        if not new_tracks:
            print(f"  [all duplicates, skipping release] {album_title}")
            continue

        print(f"  {album_title} ({release_type}) — {len(new_tracks)} new tracks")

        skip_title_keywords = {"intro", "introduction", "(intro)", "untitled", "untitled track", "untitled song", "bonus material"}

        for track_title in new_tracks:
            seen_songs.add(clean_track_title(track_title))
            if track_title.strip().lower() in skip_title_keywords:
                track_include = ""
            else:
                track_include = include_value  # inherits from release
            ws.append([track_include, album_title, year, release_type, subgenres, vocalist_gender, track_title, clean_track_title(track_title), "", "", "", ""])


    print("out of for loop")
    band_path = os.path.join(lyric_folder_path, band_name)
    os.makedirs(band_path, exist_ok=True)
    path = os.path.join(band_path, f"{band_name}_preview.xlsx")
    wb.save(path)
    print(f"Saved {path} — review and edit, then call get_band_lyrics('{band_name}')")

def load_bands_to_process(vs_folder):
    xlsx_path = os.path.join(vs_folder, "bands_to_process.xlsx")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    bands = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        band_name, country, vocalist_gender, genius_name, discogs_name, skip, OG_order, validity = row
        if not band_name or skip == "YES":
            continue
        bands.append({
            "band_name": band_name,
            "country": country,
            "vocalist_gender": vocalist_gender or "male",
            "genius_name": genius_name or band_name,
            "discogs_name": discogs_name or band_name
        })
    return bands

bands_to_process = load_bands_to_process(vs_folder)

for band in bands_to_process:
    print(f"\n{'='*50}")
    print(f"Processing: {band['band_name']}")
    print(f"{'='*50}")
    preview_discography(
        band["band_name"],
        band["vocalist_gender"],
        discogs_name=band["discogs_name"]
    )
    print("Waiting 60 seconds before starting next band preview...")
    time.sleep(60)